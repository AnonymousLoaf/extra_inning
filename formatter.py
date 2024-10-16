import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side


def auto_adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column name
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def set_fixed_column_width(sheet, width=14):
    for column in sheet.columns:
        column_letter = column[0].column_letter  # Get the column name
        sheet.column_dimensions[column_letter].width = width


def format_excel(file_path):
    # Define the groups of columns
    groups = [
        ["LastYearRank"],
        ["RankedThisYear", "WhereToRankRegionallyCoach", "WhereToRankNationallyCoach", "PrefferedRecomendation", "RankedPrevious", "PlayerPreviousRanked", "PlayerRanking"],
        ["Done"],
        [
            "PlayerHometown",
            "HS",
            "PlayerHighSchool",
            "PlayerTwitter",
            "x",
            "PlayerGPA",
            "PlayerRegion",
            "PlayerCommitted",
            "PlayerUniversity",
            "ActionVideo",
            "Committed?",
            "PlayerCommittedTo",
            "PlayerCommitted"
        ],
        ["PlayerFirstName", "PlayerLastName"],
        ["Notes"],
        [
            "Slapper",
            "PlayerPA",
            "PlayerAB",
            "PlayerBA",
            "PlayerOBP",
            "PlayerOPS",
            "PlayerHits",
            "PlayerDoubles",
            "PlayerTriples",
            "PlayerHR",
            "PlayerRBI",
            "PlayerStrikeOuts",
        ],
        [
            "CalculatedFieldingPerc",
            "FieldingPerc",
            "TotalChances",
            "Assist",
            "Putouts",
            "PlayerArmVelo",
        ],
        [
            "PlayerERA",
            "PlayerWHIP",
            "PlayerKs",
            "PlayerBB",
            "PlayerIP",
            "PlayerBAA",
            "PlayerFastballSpeed",
            "PlayerChangeUpSpeed",
        ],
        [
            "PlayerPopTime",
            "PlayerArmVelo",
            "PlayerSB",
            "PlayerATT",
            "FieldingPerc",
        ],
        ["PlayerPosition"],
        ["ClubTeamName"],
        ["GameChanger Name"],
        [
            "CoachRecommend50",
            "PlayerRankingEstimate",
        ],
        [
            "PlayerAccomplishments",
            "PlayerQuote",
        ],
        [
            "Top3Org",
            "Top Tournaments",
        ],
        [
            "Else",
        ],
        [
            "Parent1Name",
            "Parent1Email",
            "Parent1Phone",
            "Parent2Name",
            "Parent2Email",
            "Parent2Phone",
        ],
        ["CoachQuote", "PlayersRecOutsideOrg"],
        ['TopTournaments', 'WhatElse'],
        ["ParentFirstName", "ParentLastName", "ParentEmail", "ParentPhone"],
        [
            "Headshot",
        ],
        [
            "ContactFirstName",
            "ContactLastName",
            "ContactEmail",
            "ContactPhone",
        ],
        [
            "NominatorFirstName",
            "NominatorLastName",
            "NominatorEmail",
            "NominatorPhone",
        ],
        [
            "CoachNameFirst",
            "CoachNameLast",
            "CoachEmail",
            "CoachPhone",
        ],
        [
            "AthletesGoLiveName",
            "OrgLeaderFirstName",
            "OrgLeaderLastName",
        ],
        [
            "OrgLeader",
            "OrgEmail",
            "OrgPhone",
        ],
    ]

    # Define the colors for the groups
    colors = [
        "FFFFE0E0",  # Lite Red
        "FFFFE0B2",  # Lite Orange
        "FFFFFFE0",  # Lite Yellow
        "FFE0FFE0",  # Lite Green
        "FFE0FFFF",  # Lite Cyan
        "FFE0E0FF",  # Lite Blue
        "FFFFE0FF",  # Lite Purple
    ]

    # Define the colors for the headers
    header_colors = [
        "FFB1B1",  # Red
        "FFD292",  # Orange
        "FFFFB1",  # Yellow
        "BBFFBB",  # Green
        "AEFFFF",  # Cyan
        "B1B1FF",  # Blue
        "FFB1FF",  # Purple
    ]

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Define border style
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Define alignment style
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply colors to the groups and the header row
    for group_idx, group in enumerate(groups):
        color = colors[group_idx % len(colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Apply the specified header color
        header_color = header_colors[group_idx % len(header_colors)]
        header_fill = PatternFill(
            start_color=header_color, end_color=header_color, fill_type="solid"
        )

        for col_name in group:
            col_idx = None
            for cell in sheet[1]:  # Assumes first row contains headers
                if cell.value == col_name:
                    col_idx = cell.column
                    break
            if col_idx is not None:
                # Apply the specified header color to the header cell
                header_cell = sheet.cell(row=1, column=col_idx)
                header_cell.fill = header_fill
                header_cell.alignment = center_alignment
                header_cell.border = border
                
                # Apply the regular color to the rest of the cells in the column
                for cell in sheet.iter_cols(
                    min_col=col_idx, max_col=col_idx, min_row=2
                ):
                    for single_cell in cell:
                        single_cell.alignment = center_alignment
                        single_cell.border = border

    # Define the fill styles for highlighting
    light_red_fill = PatternFill(
        start_color="FFADB0", end_color="FFADB0", fill_type="solid"
    )
    dark_red_fill = PatternFill(
        start_color="FF5B61", end_color="FF5B61", fill_type="solid"
    )

    # Get the header row values as a list to match column names
    header = [cell.value for cell in sheet[1]]

    # Iterate through rows to apply conditional formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        is_red_flag_cell = row[0]  # Assuming 'is_red_flag' is the first column

        # Check if the 'is_red_flag' cell has any non-empty value
        if is_red_flag_cell.value and is_red_flag_cell.value != "[]":
            # Highlight the entire row in light red
            for cell in row:
                cell.fill = light_red_fill
                cell.alignment = center_alignment
                cell.border = border

            # Extract potential error columns from the 'is_red_flag' cell value
            error_columns = str(is_red_flag_cell.value).strip("[]").split(",")
            error_columns = [col.strip().strip("'").strip() for col in error_columns]

            # Highlight each error column in dark red
            for col in error_columns:
                if col in header:
                    col_index = header.index(
                        col
                    )  # Get the index of the error column from the header row
                    row[col_index].fill = (
                        dark_red_fill  # Apply dark red fill to the error column in the current row
                    )
                    row[col_index].alignment = center_alignment
                    row[col_index].border = border

    # Set all rows to a height of 30
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 30
    
    # Find and delete the 'is_red_flag' column
    header = [cell.value for cell in sheet[1]]
    # if "is_red_flag" in header:
    #     col_index = header.index("is_red_flag") + 1  # openpyxl is 1-indexed
    #     sheet.delete_cols(col_index)

    set_fixed_column_width(sheet)
    workbook.save(file_path)
